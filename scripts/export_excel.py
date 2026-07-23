#!/usr/bin/env python3
"""Export job postings to an Excel file with match score, status dropdown, and row coloring.

Usage:
  python3 export_excel.py --state state/seen_postings.json --config config.json
      [--output embedded_岗位追踪.xlsx]

Creates or updates an Excel file. If the file already exists, user-edited columns
(投递状态, 面经参考) are preserved; all other columns are refreshed from state data.

Key features:
  - 投递状态 column has a dropdown (data validation)
  - Selecting a status colors the ENTIRE ROW via conditional formatting (dynamic, not static)
  - No column filter dropdowns (clean header row)
"""
import argparse
import json
import os
import re
import sys
from datetime import date, datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Column definitions ──────────────────────────────────────────────
COLUMNS = [
    ("序号",        6),
    ("公司",        18),
    ("岗位名称",    30),
    ("城市",        20),
    ("岗位亮点",    36),
    ("匹配度",      10),
    ("来源链接",    30),
    ("投递入口",    30),
    ("面经参考",    30),
    ("来源平台",    16),
    ("首次发现",    12),
    ("最后确认",    12),
    ("已核实",      8),
    ("投递状态",    12),
]

HEADER_NAMES = [c[0] for c in COLUMNS]
USER_EDITABLE = {"投递状态", "面经参考"}

# ── 投递状态选项 ────────────────────────────────────────────────────
STATUS_OPTIONS = ["未投递", "已投递", "笔试中", "面试中", "已offer", "已挂", "不感兴趣"]

STATUS_MIGRATE = {
    "待投递": "未投递",
    "已笔试": "笔试中",
    "已面试": "面试中",
}

# ── Row colors for each status (used in conditional formatting) ─────
# These are applied to the ENTIRE ROW dynamically when user selects a status
STATUS_COLORS = {
    "未投递":       None,        # 无填充（白色默认）
    "已投递":       "FFF2CC",    # 浅黄
    "笔试中":       "DDEBF7",    # 浅蓝
    "面试中":       "FCE4D6",    # 浅粉
    "已offer":      "C6EFCE",    # 浅绿
    "已挂":         "FFC7CE",    # 浅红
    "不感兴趣":     "E0E0E0",    # 灰色
}

# ── Styling ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
CELL_FONT   = Font(name="微软雅黑", size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SCORE_FILLS = {
    5: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    4: PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    3: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    2: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    1: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

DEADLINE_URGENT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
DEADLINE_SOON   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ── Helpers ─────────────────────────────────────────────────────────

def load_json(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def calc_match_score(title, highlight, positive_keywords, fuzzy_keywords):
    text = f"{title} {highlight}".lower()
    pos_hits = sum(1 for kw in positive_keywords if kw.lower() in text)
    fuzzy_hits = sum(1 for kw in fuzzy_keywords if kw.lower() in text)
    if pos_hits >= 3:
        return 5, "⭐⭐⭐⭐⭐"
    elif pos_hits == 2:
        return 4, "⭐⭐⭐⭐"
    elif pos_hits == 1:
        return 3, "⭐⭐⭐"
    elif fuzzy_hits >= 1:
        return 2, "⭐⭐"
    else:
        return 1, "⭐"


def extract_deadline(highlight, title):
    text = f"{highlight} {title}"
    patterns = [
        r"截止(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})",
        r"截止\s*(\d{1,2})月(\d{1,2})日",
        r"截止\s*(\d{1,2})[-/](\d{1,2})",
        r"deadline[:\s]*(\d{4})[-/](\d{1,2})[-/](\d{1,2})",
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            groups = m.groups()
            if len(groups) == 3:
                return f"{groups[0]}-{int(groups[1]):02d}-{int(groups[2]):02d}"
            elif len(groups) == 2:
                year = date.today().year
                return f"{year}-{int(groups[0]):02d}-{int(groups[1]):02d}"
    return ""


def days_until(deadline_str):
    try:
        d = date.fromisoformat(deadline_str)
        delta = (d - date.today()).days
        return delta if delta >= 0 else -1
    except (ValueError, TypeError):
        return None


def nowcoder_interview_url(company):
    import urllib.parse
    return f"https://www.nowcoder.com/search?type=post&query={urllib.parse.quote(company + ' 面经')}"


def row_key(company, title):
    return f"{company}|{title}"


# ── Main export logic ───────────────────────────────────────────────

def build_row_data(rec, config, idx):
    title = rec.get("title", "")
    highlight = rec.get("highlight", "")
    pos_kw = config.get("job_filter", {}).get("positive_keywords", [])
    fz_kw = config.get("job_filter", {}).get("fuzzy_keywords", [])

    score_val, score_str = calc_match_score(title, highlight, pos_kw, fz_kw)
    apply_url = rec.get("apply_url", "")
    interview_url = rec.get("interview_url", "") or nowcoder_interview_url(rec.get("company", ""))

    return {
        "序号": idx,
        "公司": rec.get("company", ""),
        "岗位名称": title,
        "城市": rec.get("city", "未注明"),
        "岗位亮点": highlight,
        "匹配度": score_str,
        "来源链接": rec.get("source_url", ""),
        "投递入口": apply_url,
        "面经参考": interview_url,
        "来源平台": rec.get("source_platform", ""),
        "首次发现": rec.get("first_seen", ""),
        "最后确认": rec.get("last_confirmed", ""),
        "已核实": "✅" if rec.get("verified") else "—",
        "投递状态": "未投递",
        "_score_val": score_val,
    }


def load_existing_user_data(xlsx_path):
    if not os.path.exists(xlsx_path):
        return {}
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_map = {h: i for i, h in enumerate(headers) if h}

        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            company = row[col_map.get("公司", 1)] if "公司" in col_map and col_map["公司"] < len(row) else ""
            title = row[col_map.get("岗位名称", 2)] if "岗位名称" in col_map and col_map["岗位名称"] < len(row) else ""
            if not company or not title:
                continue
            key = row_key(company, title)
            user_data = {}
            for col_name in USER_EDITABLE:
                if col_name in col_map and col_map[col_name] < len(row):
                    val = row[col_map[col_name]]
                    if val and str(val).strip() and str(val).strip() != nowcoder_interview_url(company):
                        user_data[col_name] = str(val).strip()
            if user_data:
                result[key] = user_data
        wb.close()
        return result
    except Exception:
        return {}


def write_excel(xlsx_path, rows, config):
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位总览"

    n_cols = len(COLUMNS)
    n_rows = len(rows)
    last_col_letter = get_column_letter(n_cols)
    last_row = n_rows + 1
    # 投递状态 column letter (for conditional formatting formulas)
    status_col_idx = [i + 1 for i, (name, _) in enumerate(COLUMNS) if name == "投递状态"][0]
    status_col_letter = get_column_letter(status_col_idx)

    # ── Header row ──
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ──
    for row_idx, row_data in enumerate(rows, 2):
        score_val = row_data.pop("_score_val", 1)
        for col_idx, (name, _) in enumerate(COLUMNS, 1):
            val = row_data.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

            # Score color (static — this doesn't change)
            if name == "匹配度":
                cell.fill = SCORE_FILLS.get(score_val, SCORE_FILLS[1])
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 已核实 column styling
            if name == "已核实":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "✅":
                    cell.font = Font(name="微软雅黑", size=10, color="008000", bold=True)
                else:
                    cell.font = Font(name="微软雅黑", size=10, color="999999")

            # 投递状态 cell styling (bold + centered, color via conditional formatting)
            if name == "投递状态":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="微软雅黑", size=10, bold=True)

            # Hyperlinks
            if name in ("来源链接", "投递入口", "面经参考") and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    # ── Data validation: 投递状态 dropdown ──
    dv = DataValidation(
        type="list",
        formula1='"未投递,已投递,笔试中,面试中,已offer,已挂,不感兴趣"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "请从下拉列表中选择投递状态"
    dv.errorTitle = "无效输入"
    dv.prompt = "点击下拉箭头选择投递状态"
    dv.promptTitle = "投递状态"
    dv.add(f"{status_col_letter}2:{status_col_letter}{max(last_row, 200)}")
    ws.add_data_validation(dv)

    # ── Conditional formatting: entire row colors based on 投递状态 ──
    # Formula references the status column with absolute column, relative row
    # e.g. =$N2="已投递"  →  colors the entire row when status cell = "已投递"
    cf_range = f"A2:{last_col_letter}{max(last_row, 200)}"

    for status, color in STATUS_COLORS.items():
        if color is None:
            continue
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        formula = f'${status_col_letter}2="{status}"'
        rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=False)
        ws.conditional_formatting.add(cf_range, rule)

    # ── No auto-filter (user doesn't want column filter buttons) ──
    # ws.auto_filter.ref = None  # explicitly no filter

    # Freeze header + first 3 columns (序号/公司/岗位)
    ws.freeze_panes = "D2"

    # Row height
    ws.row_dimensions[1].height = 28
    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = 36

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Export job postings to Excel")
    ap.add_argument("--state", required=True, help="Path to seen_postings.json")
    ap.add_argument("--config", required=True, help="Path to config.json")
    ap.add_argument("--output", help="Output Excel file path (auto-derived if omitted)")
    args = ap.parse_args()

    config = load_json(args.config, {})
    state = load_json(args.state, {"postings": {}})
    postings = state.get("postings", {})

    if not postings:
        print("WARNING: No postings in state file, Excel will be empty.", file=sys.stderr)

    category = config.get("job_category_label", "岗位")
    if not args.output:
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(args.state)))
        args.output = os.path.join(project_dir, f"{category}_岗位追踪.xlsx")

    existing_user = load_existing_user_data(args.output)

    rows = []
    sorted_postings = sorted(postings.values(), key=lambda r: r.get("first_seen", ""), reverse=True)
    for idx, rec in enumerate(sorted_postings, 1):
        row = build_row_data(rec, config, idx)
        key = row_key(row["公司"], row["岗位名称"])
        if key in existing_user:
            for col_name in USER_EDITABLE:
                if col_name in existing_user[key]:
                    val = existing_user[key][col_name]
                    if col_name == "投递状态" and val in STATUS_MIGRATE:
                        val = STATUS_MIGRATE[val]
                    row[col_name] = val
        rows.append(row)

    write_excel(args.output, rows, config)

    n_total = len(rows)
    n_apply = sum(1 for r in rows if r.get("投递入口"))
    print(f"Excel exported: {args.output}")
    print(f"  Total postings: {n_total}")
    print(f"  With apply URL: {n_apply}")
    print(f"  User data preserved: {len(existing_user)} rows")


if __name__ == "__main__":
    main()
